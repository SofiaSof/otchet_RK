import subprocess
import time
import json
import os
import re
import requests
import websocket

import openpyxl


def get_page_target_ws():
    chrome_proc = subprocess.Popen([
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        "--headless=new",
        "--remote-debugging-port=9222",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--disable-gpu",
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    for _ in range(30):
        try:
            resp = requests.get("http://localhost:9222/json", timeout=5)
            for t in resp.json():
                if t["type"] == "page":
                    ws_url = t["webSocketDebuggerUrl"]
                    ws = websocket.create_connection(ws_url, suppress_origin=True, timeout=30)
                    return ws, chrome_proc
        except Exception:
            time.sleep(0.5)

    chrome_proc.terminate()
    raise RuntimeError("Could not connect to Chrome CDP")


def navigate_and_wait(ws, url, wait_seconds=8):
    ws.send(json.dumps({"id": 1, "method": "Page.navigate", "params": {"url": url}}))
    ws.recv()
    time.sleep(wait_seconds)

    ws.send(json.dumps({"id": 2, "method": "Runtime.evaluate",
                         "params": {"expression": "document.readyState", "returnByValue": True}}))
    resp = json.loads(ws.recv())
    state = resp.get("result", {}).get("result", {}).get("value")
    if state != "complete":
        for _ in range(20):
            time.sleep(1)
            ws.send(json.dumps({"id": 3, "method": "Runtime.evaluate",
                                 "params": {"expression": "document.readyState", "returnByValue": True}}))
            resp = json.loads(ws.recv())
            state = resp.get("result", {}).get("result", {}).get("value")
            if state == "complete":
                break


def find_photo_urls(ws, photo_type="all"):
    script = """
    (function() {
        var results = [];
        var seen = {};
        var photoType = arguments[0];
        
        document.querySelectorAll("img").forEach(function(img) {
            var src = img.src || "";
            var isPhoto = (
                src.includes("onlineapi.russoutdoor.ru") ||
                src.includes("russoutdoor.ru/photo") ||
                src.includes("CampaignGuaranteedPhotoReport") ||
                src.includes("GetPosterContent") ||
                (img.className && img.className.includes("thumbnail"))
            );
            
            var isStreetBanner = src.includes("GetPosterContent") || src.includes("CampaignGuaranteedPhotoReport");
            var isStreet = img.alt && (img.alt.toLowerCase().includes("улиц") || img.alt.toLowerCase().includes("баннер"));
            var isStreetContainer = img.closest && img.closest('[class*="street"], [class*="banner"], [class*="outdoor"]');
            
            var includePhoto = false;
            if (photoType === "street_banner") {
                includePhoto = isPhoto && (isStreetBanner || isStreet || isStreetContainer);
            } else {
                includePhoto = isPhoto;
            }
            
            if (includePhoto && src && !src.includes(".svg") && !seen[src]) {
                seen[src] = true;
                results.push(src);
            }
        });
        return JSON.stringify(results);
    })()
    """
    ws.send(json.dumps({"id": 10, "method": "Runtime.evaluate",
                         "params": {"expression": script, "returnByValue": True, "arguments": [{"value": photo_type}]}}))
    resp = json.loads(ws.recv())
    return json.loads(resp.get("result", {}).get("result", {}).get("value", "[]"))


def download_image(url, folder, session):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://online.russoutdoor.ru/"
    }
    try:
        r = session.get(url, timeout=30, headers=headers, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 1000:
            ext = ".jpg"
            if "png" in r.headers.get("content-type", "") or url.endswith(".png"):
                ext = ".png"
            elif "webp" in r.headers.get("content-type", "") or url.endswith(".webp"):
                ext = ".webp"

            base_name = "preview"
            filepath = os.path.join(folder, f"{base_name}{ext}")
            
            counter = 1
            while os.path.exists(filepath):
                filepath = os.path.join(folder, f"{base_name}_{counter}{ext}")
                counter += 1
            
            with open(filepath, "wb") as f:
                f.write(r.content)
            return filepath
    except Exception as e:
        print(f"    Download error: {e}")
    return None


def sanitize_folder(name):
    name = re.sub(r'[<>:"/\\|?*]', '_', str(name))
    return name.strip() or "unnamed"


def main():
    excel_path = "C:/test/Otchet_Samocat.xlsx"
    row_nums = list(range(21, 104))

    wb = openpyxl.load_workbook(excel_path)
    ws_excel = wb.active

    session = requests.Session()
    ws_cdp = None
    chrome_proc = None

    try:
        ws_cdp, chrome_proc = get_page_target_ws()

        for row_num in row_nums:
            print(f"\n{'='*50}")
            print(f"Processing row {row_num}")
            print('='*50)

            folder_name = sanitize_folder(ws_excel.cell(row=row_num, column=11).value or "")
            city = sanitize_folder(ws_excel.cell(row=row_num, column=22).value or "")
            if not folder_name or folder_name == "unnamed":
                print(f"Row {row_num}: column K is empty")
                continue
            if not city or city == "unnamed":
                city = "unknown"

            url = None
            cell_s = ws_excel.cell(row=row_num, column=19)
            if cell_s.hyperlink:
                url = cell_s.hyperlink.target
            if not url:
                url = cell_s.value
            if not url:
                print(f"Row {row_num}: column S has no URL")
                continue

            print(f"URL: {url}")
            print(f"Folder: {folder_name}/{city}")

            output_folder = os.path.join("C:/test/photos", folder_name, city)
            os.makedirs(output_folder, exist_ok=True)

            navigate_and_wait(ws_cdp, url, wait_seconds=6)
            photo_urls = find_photo_urls(ws_cdp, photo_type="street_banner")

            print(f"\nFound {len(photo_urls)} photo(s):")
            for u in photo_urls:
                print(f"  {u}")

            downloaded = 0
            for i, photo_url in enumerate(photo_urls, 1):
                if "preview" not in photo_url.lower():
                    print(f"  [{i}] Skipped (not preview): {photo_url}")
                    continue
                filepath = download_image(photo_url, output_folder, session)
                if filepath:
                    print(f"  [{i}] Saved: {os.path.basename(filepath)} ({len(open(filepath,'rb').read())} bytes)")
                    downloaded += 1
                else:
                    print(f"  [{i}] Failed: {photo_url}")

            print(f"\nRow {row_num}: Downloaded {downloaded}/{len(photo_urls)} photos to:\n{output_folder}")

    finally:
        if ws_cdp:
            ws_cdp.close()
        if chrome_proc:
            chrome_proc.terminate()
            chrome_proc.wait()


if __name__ == "__main__":
    main()
